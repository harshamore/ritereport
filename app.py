import streamlit as st
import pandas as pd
import json
import os
from io import BytesIO
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Financial Reporting MVP", layout="centered")

st.title("Automated Financial Reporting (MVP)")

MAPPING_FILE = "mapping.json"

if "chart_mapping" not in st.session_state:
    if os.path.exists(MAPPING_FILE):
        with open(MAPPING_FILE, "r") as f:
            st.session_state.chart_mapping = json.load(f)
    else:
        st.session_state.chart_mapping = {}

categories = [
    "Current Assets",
    "Long-term Assets",
    "Current Liabilities",
    "Long-term Liabilities",
    "Equity",
    "Revenue",
    "Operating Expenses",
    "Non-Operating Expenses"
]

uploaded_file = st.file_uploader("Upload a file of your trial balance (CSV or Excel)", type=["csv", "xlsx", "xls"])

def save_mappings():
    """Save the current chart mappings to a JSON file."""
    with open(MAPPING_FILE, "w") as f:
        json.dump(st.session_state.chart_mapping, f)

def read_trial_balance(file):
    """Read the trial balance from CSV or Excel."""
    if file.name.endswith(".csv"):
        df = pd.read_csv(file)
    else:
        df = pd.read_excel(file)
    return df

def create_pdf(income_df, bs_df):
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, "Income Statement", ln=True)
    pdf.set_font("Arial", "", 12)
    for _, row in income_df.iterrows():
        pdf.cell(0, 10, f"{row['Description']}: ${row['Amount']:,.2f}", ln=True)

    pdf.ln(10)
    pdf.set_font("Arial", "B", 16)
    pdf.cell(0, 10, "Balance Sheet", ln=True)
    pdf.set_font("Arial", "", 12)
    for _, row in bs_df.iterrows():
        pdf.cell(0, 10, f"{row['Description']}: ${row['Amount']:,.2f}", ln=True)

    pdf_output = BytesIO()
    pdf.output(pdf_output, "F")
    pdf_output.seek(0)
    return pdf_output

def create_excel_with_hyperlinks(trial_df, income_df, bs_df, category_first_account_map):
    """Create an Excel file with Trial Balance, Income Statement, and Balance Sheet.
    Income Statement and Balance Sheet cells link to corresponding accounts in the Trial Balance."""

    wb = Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # 1. Trial Balance Sheet
    tb_sheet = wb.create_sheet("Trial Balance")
    tb_cols = list(trial_df.columns)
    # Write headers
    for col_idx, col_name in enumerate(tb_cols, start=1):
        tb_sheet.cell(row=1, column=col_idx).value = col_name
    
    # Write rows
    for row_idx, row in enumerate(trial_df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            tb_sheet.cell(row=row_idx, column=col_idx).value = value

    # We have category_first_account_map which maps category -> (row_number_of_first_account)
    # We'll use this to create hyperlinks in the IS and BS.

    # 2. Income Statement Sheet
    is_sheet = wb.create_sheet("Income Statement")
    is_cols = ["Description", "Amount"]
    for col_idx, col_name in enumerate(is_cols, start=1):
        is_sheet.cell(row=1, column=col_idx).value = col_name

    for row_idx, (desc, amt) in enumerate(zip(income_df["Description"], income_df["Amount"]), start=2):
        is_sheet.cell(row=row_idx, column=1).value = desc
        is_sheet.cell(row=row_idx, column=2).value = amt

        # Attempt to link to the first account of the corresponding category (if applicable)
        # Map descriptions to categories:
        desc_to_category_map = {
            "Revenue": "Revenue",
            "Operating Expenses": "Operating Expenses",
            "Non-Operating Expenses": "Non-Operating Expenses",
            "Net Income": None  # Net Income is derived, not directly linked to accounts
        }
        cat = desc_to_category_map.get(desc, None)
        if cat and cat in category_first_account_map:
            row_num = category_first_account_map[cat]
            # Create a hyperlink in the Description cell linking to trial balance row
            # Hyperlinks in openpyxl format: cell.hyperlink = "#'SheetName'!A1"
            is_sheet.cell(row=row_idx, column=1).hyperlink = f"#{'Trial Balance'}!A{row_num}"
            is_sheet.cell(row=row_idx, column=1).style = "Hyperlink"

    # 3. Balance Sheet Sheet
    bs_sheet = wb.create_sheet("Balance Sheet")
    bs_cols = ["Description", "Amount"]
    for col_idx, col_name in enumerate(bs_cols, start=1):
        bs_sheet.cell(row=1, column=col_idx).value = col_name

    for row_idx, (desc, amt) in enumerate(zip(bs_df["Description"], bs_df["Amount"]), start=2):
        bs_sheet.cell(row=row_idx, column=1).value = desc
        bs_sheet.cell(row=row_idx, column=2).value = amt

        # Attempt to find a category that matches the description
        # We'll map descriptions back to categories
        desc_to_cat_for_bs = {
            "Current Assets": "Current Assets",
            "Long-term Assets": "Long-term Assets",
            "Total Assets": None,
            "Current Liabilities": "Current Liabilities",
            "Long-term Liabilities": "Long-term Liabilities",
            "Equity": "Equity",
            "Total Liabilities & Equity": None
        }
        cat = desc_to_cat_for_bs.get(desc, None)
        if cat and cat in category_first_account_map:
            row_num = category_first_account_map[cat]
            bs_sheet.cell(row=row_idx, column=1).hyperlink = f"#{'Trial Balance'}!A{row_num}"
            bs_sheet.cell(row=row_idx, column=1).style = "Hyperlink"

    # Save to BytesIO
    excel_bytes = BytesIO()
    wb.save(excel_bytes)
    excel_bytes.seek(0)
    return excel_bytes

if uploaded_file is not None:
    df = read_trial_balance(uploaded_file)
    required_columns = {"Account Name", "Debit", "Credit"}
    if not required_columns.issubset(df.columns):
        st.error("The file must contain columns: Account Name, Debit, Credit")
    else:
        total_debits = df["Debit"].sum()
        total_credits = df["Credit"].sum()
        if abs(total_debits - total_credits) > 1e-6:
            st.warning(f"Total Debits (${total_debits:,.2f}) do not equal Total Credits (${total_credits:,.2f}). Please check your data.")
        else:
            st.success("Trial balance debits and credits are balanced.")

        st.subheader("Uploaded Trial Balance")
        st.dataframe(df)

        # Map accounts form
        account_names = df["Account Name"].unique()

        st.subheader("Map Accounts to Categories")
        with st.form("mapping_form"):
            for account in account_names:
                default_cat = st.session_state.chart_mapping.get(account, categories[0])
                cat = st.selectbox(f"{account}:", categories, index=categories.index(default_cat) if default_cat in categories else 0)
                st.session_state.chart_mapping[account] = cat
            submitted = st.form_submit_button("Save Mapping")

        if submitted:
            save_mappings()
            st.success("Mappings saved.")

        if st.button("Generate Financial Statements"):
            df["Category"] = df["Account Name"].map(st.session_state.chart_mapping)
            df["Amount"] = df["Debit"] - df["Credit"]

            cat_sums = df.groupby("Category")["Amount"].sum()

            # Calculate BS and IS components
            total_current_assets = cat_sums.get("Current Assets", 0)
            total_long_assets = cat_sums.get("Long-term Assets", 0)
            total_assets = total_current_assets + total_long_assets

            total_current_liab = cat_sums.get("Current Liabilities", 0)
            total_long_liab = cat_sums.get("Long-term Liabilities", 0)
            total_liabilities = total_current_liab + total_long_liab
            total_equity = cat_sums.get("Equity", 0)

            total_revenue = cat_sums.get("Revenue", 0)
            total_op_exp = cat_sums.get("Operating Expenses", 0)
            total_non_op_exp = cat_sums.get("Non-Operating Expenses", 0)
            total_expenses = total_op_exp + total_non_op_exp
            net_income = total_revenue - total_expenses

            # Display Income Statement
            st.subheader("Income Statement")
            st.write("**Revenue**:", f"${total_revenue:,.2f}")
            st.write("**Operating Expenses**:", f"${total_op_exp:,.2f}")
            st.write("**Non-Operating Expenses**:", f"${total_non_op_exp:,.2f}")
            st.write("---")
            st.write("**Net Income**:", f"${net_income:,.2f}")

            # Display Balance Sheet
            st.subheader("Balance Sheet")
            col1, col2 = st.columns(2)
            with col1:
                st.write("**Assets**")
                st.write("Current Assets:", f"${total_current_assets:,.2f}")
                st.write("Long-term Assets:", f"${total_long_assets:,.2f}")
                st.write("---")
                st.write("Total Assets:", f"${total_assets:,.2f}")
            with col2:
                st.write("**Liabilities & Equity**")
                st.write("Current Liabilities:", f"${total_current_liab:,.2f}")
                st.write("Long-term Liabilities:", f"${total_long_liab:,.2f}")
                st.write("Equity:", f"${total_equity:,.2f}")
                st.write("---")
                st.write("Total L&E:", f"${(total_liabilities + total_equity):,.2f}")

            if abs(total_assets - (total_liabilities + total_equity)) < 1e-6:
                st.success("Balance Sheet is balanced.")
            else:
                st.warning("Balance Sheet does not balance. Check your mappings or data.")

            # Prepare dataframes
            income_statement_df = pd.DataFrame({
                "Description": ["Revenue", "Operating Expenses", "Non-Operating Expenses", "Net Income"],
                "Amount": [total_revenue, total_op_exp, total_non_op_exp, net_income]
            })

            balance_sheet_df = pd.DataFrame({
                "Description": ["Current Assets", "Long-term Assets", "Total Assets", 
                                "Current Liabilities", "Long-term Liabilities", "Equity", "Total Liabilities & Equity"],
                "Amount": [total_current_assets, total_long_assets, total_assets,
                           total_current_liab, total_long_liab, total_equity, total_liabilities + total_equity]
            })

            # For hyperlinking, find the first account row for each category in the trial balance df
            # We'll include the trial balance with category and amount info for clarity
            trial_df = df[["Account Name", "Debit", "Credit", "Category", "Amount"]].copy().reset_index(drop=True)
            # trial_df rows start at 2 in Excel (since row 1 is header)
            category_first_account_map = {}
            for i, row in trial_df.iterrows():
                cat = row["Category"]
                if cat not in category_first_account_map:
                    # Save the Excel row number for the first occurrence of this category
                    category_first_account_map[cat] = i+2  # +2 because Excel headers start at row 1, data at row 2

            # Create Excel with hyperlinks
            excel_file = create_excel_with_hyperlinks(trial_df, income_statement_df, balance_sheet_df, category_first_account_map)

            # Download buttons for Excel
            st.subheader("Export Reports")
            st.download_button(
                label="Download All Reports (Excel)",
                data=excel_file,
                file_name="financial_reports.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
